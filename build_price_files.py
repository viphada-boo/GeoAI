# -*- coding: utf-8 -*-
"""สร้าง Gistda_Price_List.csv / .xlsx / .html จากข้อมูลที่ extract จาก Gistda_Price_List.pdf"""
import csv
import html as html_mod

import pandas as pd

COLS_STD = ["ดาวเทียม (Satellite)", "รายละเอียดภาพ (Resolution)",
            "ข้อมูลในคลัง (Standard Archive)", "ข้อมูลชนิดสั่งถ่าย (Standard Tasking)"]
COLS_MODE = ["Mode", "รายละเอียดภาพ (Resolution)",
             "ข้อมูลในคลัง (Standard Archive)", "ข้อมูลชนิดสั่งถ่าย (Standard Tasking)"]

SECTIONS = [
    {
        "sheet": "1-สูงมาก 30-50 ซม.",
        "title": "ราคาข้อมูลจากดาวเทียมรายละเอียดสูงมาก (รายละเอียดภาพ 30 – 50 ซม.)",
        "unit": "หน่วย : บาท/ตร.กม.",
        "columns": COLS_STD,
        "rows": [
            ["Pléiades NEO", "30 cm.", "880", "1,270"],
            ["WorldView-4", "30 cm.", "920", "1,560"],
            ["SuperView-2", "42 cm.", "700", "1,100"],
            ["WorldView-1", "50 cm.", "700", "1,100"],
            ["WorldView-2", "50 cm.", "700", "1,100"],
            ["WorldView-3", "50 cm.", "700", "1,100"],
            ["GeoEye-1", "50 cm.", "700", "1,100"],
            ["Pléiades", "50 cm.", "490", "830"],
            ["EarthScanner", "50 cm.", "400", "800"],
            ["SuperView-1", "50 cm.", "500", "900"],
            ["KOMPSAT-3", "50 cm.", "400", "700"],
            ["SKYSAT", "50 cm.", "300", "560"],
        ],
        "notes": [
            "ข้อมูลในคลัง (Standard Archive) พื้นที่การสั่งขั้นต่ำ 25 ตร.กม. / ข้อมูลชนิดสั่งถ่าย (Standard Tasking) พื้นที่การสั่งขั้นต่ำ 100 ตร.กม.",
            "ราคาข้างต้นใช้สำหรับ level Primary (PAN, MS, Pansharpened)",
            "SKYSAT: ข้อมูลในคลังพื้นที่การสั่งขั้นต่ำ 1,250 ตร.กม. เข้าดูผ่าน API/Explorer / ข้อมูลสั่งถ่ายโปรดติดต่อเจ้าหน้าที่",
        ],
    },
    {
        "sheet": "2-สูง 60ซม.-2ม.",
        "title": "ราคาข้อมูลจากดาวเทียมรายละเอียดสูง (รายละเอียดภาพ 60 ซม. – 2 ม.)",
        "unit": "หน่วย : บาท/ตร.กม.",
        "columns": COLS_STD,
        "rows": [
            ["QuickBird", "60 cm.", "700", "N/A"],
            ["GaoFen-7", "65 cm.", "400", "700"],
            ["Jilin", "75 cm.", "300", "600"],
            ["DailyVision", "75 cm.", "300", "600"],
            ["GaoFen-2", "80 cm.", "300", "400"],
            ["IKONOS", "1 m.", "400", "N/A"],
        ],
        "notes": [
            "ข้อมูลในคลัง (Standard Archive) พื้นที่การสั่งขั้นต่ำ 25 ตร.กม. / ข้อมูลชนิดสั่งถ่าย (Standard Tasking) พื้นที่การสั่งขั้นต่ำ 100 ตร.กม.",
            "ราคาข้างต้นใช้สำหรับ level Primary (PAN, MS, Pansharpened)",
        ],
    },
    {
        "sheet": "3-Video-Night",
        "title": "ราคาข้อมูลจากดาวเทียมรายละเอียดสูง — Video Constellation / Night Imaging",
        "unit": "หน่วย : ตามระบุในตาราง",
        "columns": COLS_STD + ["หน่วยราคา"],
        "rows": [
            ["Video Constellation", "1 m.", "142,500", "285,000", "บาท/30 วินาที"],
            ["Night Imaging", "1 m.", "800", "1,400", "บาท/ตร.กม."],
        ],
        "notes": [
            "ดาวเทียมสามารถถ่ายภาพเคลื่อนไหวในรูปแบบวิดีโอและภาพกลางคืน ภาพวิดีโอแต่ละช่วงจำกัดความยาวอยู่ที่ 30 วินาที",
            "พื้นที่การสั่งซื้อข้อมูลขั้นต่ำ ทั้งข้อมูลในคลังและข้อมูลสั่งถ่ายใหม่ เท่ากับ 100 ตร.กม.",
            "ความกว้างของแนวถ่ายภาพ อย่างน้อย 5 กม.",
        ],
    },
    {
        "sheet": "4-SPOT",
        "title": "ราคาข้อมูลจากดาวเทียมรายละเอียดสูง — SPOT",
        "unit": "หน่วย : บาท/ตร.กม.",
        "columns": COLS_STD,
        "rows": [
            ["SPOT-6", "1.5 m.", "190", "230"],
            ["SPOT-7", "1.5 m.", "190", "230"],
        ],
        "notes": [
            "ข้อมูลในคลัง (Standard Archive) พื้นที่การสั่งขั้นต่ำ 100 ตร.กม. / ข้อมูลชนิดสั่งถ่าย (Standard Tasking) พื้นที่การสั่งขั้นต่ำ 500 ตร.กม.",
        ],
    },
    {
        "sheet": "5-ไทยโชต",
        "title": "ราคาข้อมูลจากดาวเทียมรายละเอียดสูง — ไทยโชต (Thaichote)",
        "unit": "หน่วย : บาท/ภาพ",
        "columns": COLS_STD,
        "rows": [
            ["ไทยโชต", "2 m.", "700", "6,500"],
        ],
        "notes": [
            "ราคาที่ปรับ Orthorectification แล้ว ราคา 910 บาท/ภาพ",
        ],
    },
    {
        "sheet": "6-LANDSAT",
        "title": "ราคาข้อมูลจากดาวเทียมรายละเอียดปานกลาง (รายละเอียดมากกว่า 2 เมตร) — LANDSAT",
        "unit": "หน่วย : บาท/ภาพ",
        "columns": COLS_STD,
        "rows": [
            ["LANDSAT-5", "30 m.", "150", "N/A"],
            ["LANDSAT-7", "30 m.", "150", "N/A"],
            ["LANDSAT-8", "30 m.", "150", "N/A"],
            ["LANDSAT-9", "30 m.", "150", "N/A"],
        ],
        "notes": [
            "คิดเฉพาะค่าดำเนินการผลิตข้อมูลจากคลังข้อมูล",
            "LANDSAT 5 ราคาข้างต้นใช้สำหรับ Level 1T มีทั้งหมด 7 Bands / LANDSAT 7 ใช้สำหรับ Level 1T มีทั้งหมด 8 Bands",
            "LANDSAT 8, 9: หากประสงค์ให้ สทอภ. ดำเนินการดาวน์โหลด คิดค่าดำเนินการ 150 บาท/ภาพ (Level 1T มีทั้งหมด 11 Bands)",
        ],
    },
    {
        "sheet": "7-PLANETSCOPE",
        "title": "ราคาข้อมูลจากดาวเทียมรายละเอียดปานกลาง — PLANETSCOPE",
        "unit": "หน่วย : บาท/ตร.กม./ปี",
        "columns": ["ดาวเทียม (Satellite)", "รายละเอียดภาพ (Resolution)",
                    "ข้อมูลในคลัง (Standard Archive) Access+Download", "การติดตาม (Monitoring)"],
        "rows": [
            ["PLANETSCOPE", "3 m.", "180", "240"],
        ],
        "notes": [
            "พื้นที่การสั่งขั้นต่ำ 100 ตร.กม.",
            "**ระยะเวลาสัญญา 1 ปี",
            "เข้าดูข้อมูลและดาวน์โหลดผ่าน Planet Explorer, Planet API, Desktop GIS",
        ],
    },
    {
        "sheet": "8-RADARSAT-2",
        "title": "ราคาข้อมูลจากดาวเทียมระบบเรดาร์ — RADARSAT-2 (C band)",
        "unit": "หน่วย : บาท/ภาพ",
        "columns": ["Mode", "รายละเอียดภาพ (Resolution)", "Single Look complex (บาท)", "Path Image (บาท)"],
        "rows": [
            ["Standard", "25 m.", "57,600", "57,600"],
            ["Spotlight A", "1 m.", "134,400", "134,400"],
            ["Utra-Fine", "3 m.", "86,400", "86,400"],
            ["Wide Utra-Fine", "3 m.", "124,800", "124,800"],
            ["Multi-Look Fine", "8 m.", "67,200", "67,200"],
            ["Wide Multi-Look Fine", "8 m.", "120,000", "120,000"],
            ["Fine", "8 m.", "57,600", "57,600"],
            ["Wide", "30 m.", "57,600", "57,600"],
            ["ScanSAR Narrow", "50 m.", "N/A", "57,600"],
            ["ScanSAR Wide", "100 m.", "N/A", "57,600"],
            ["Extended High, Low", "25 m.", "57,600", "57,600"],
            ["Fine Quad-Pol", "8 m.", "86,400", "N/A"],
            ["Wide Fine Quad-Pol", "8 m.", "124,800", "N/A"],
        ],
        "notes": [],
    },
    {
        "sheet": "9-TerraSAR-X",
        "title": "ราคาข้อมูลจากดาวเทียมระบบเรดาร์ — TerraSAR-X (X band)",
        "unit": "หน่วย : บาท/ภาพ",
        "columns": COLS_MODE,
        "rows": [
            ["Staring Spotlight (ST)", "0.25 m.", "162,630", "325,260"],
            ["High Res Spotlight (HS)", "1 m.", "139,230", "278,460"],
            ["Spotlight", "2 m.", "99,450", "198,900"],
            ["StripMap", "3 m.", "69,030", "138,060"],
            ["ScanSAR", "18.5 m.", "40,950", "81,900"],
            ["Wide ScanSAR", "40 m.", "40,950", "81,900"],
        ],
        "notes": [],
    },
    {
        "sheet": "10-COSMO-SkyMed",
        "title": "ราคาข้อมูลจากดาวเทียมระบบเรดาร์ — COSMO SkyMed (X band)",
        "unit": "หน่วย : บาท/ภาพ",
        "columns": ["Mode", "รายละเอียดภาพ (Resolution)", "Polarization", "New Acquisition"],
        "rows": [
            ["Spotlight-2", "1x1 m.", "HH, VV", "180,000"],
            ["StripMap Himage", "3 x 3 – 5 x 5 m.", "HH, HV, VH, VV", "93,000"],
            ["StripMap PingPong", "10 x 12 – 20 x 20 m.",
             "2 ช่องสัญญาณ polarimattric : HH,VV หรือ HH,HV หรือ VV,VH", "68,000"],
            ["ScanSAR Wide", "14 x 22 – 30 x 30 m.", "HH, HV, VH, VV", "78,000"],
            ["ScanSAR Huge", "14 x 38 – 100 x 100 m.", "HH, HV, VH, VV", "78,000"],
        ],
        "notes": [],
    },
    {
        "sheet": "11-GaoFen-3",
        "title": "ราคาข้อมูลจากดาวเทียมระบบเรดาร์ — GaoFen-3 (C band)",
        "unit": "หน่วย : บาท/ภาพ",
        "columns": ["Mode", "รายละเอียดภาพ (Resolution)", "Polarization",
                    "ข้อมูลในคลัง (Standard Archive)", "ข้อมูลชนิดสั่งถ่าย (Standard Tasking)"],
        "rows": [
            ["Spotlight (SL)", "1 m.", "HH, VV", "116,400", "180,500"],
            ["Ultra-fine Stripmap (UFS)", "3 m.", "HH, VV", "68,900", "118,800"],
            ["Fine Stripmap (FSI)", "5 m.", "HH, VV", "64,200", "95,000"],
            ["Wide Fine Stripmap (FSII)", "10 m.", "HH, HV / VV, VH", "64,200", "90,300"],
            ["Standard Stripmap (SS)", "25 m.", "HH, HV / VV, VH", "54,700", "85,500"],
            ["Narrow ScanSAR (NSC)", "50 m.", "HH, HV / VV, VH", "32,100", "42,800"],
            ["Wide ScanSAR (WSC)", "100 m.", "HH, HV / VV, VH", "32,100", "45,800"],
            ["Quad-pol Stripmap (QPSI)", "8 m.", "HH, HV / VV, VH", "71,300", "137,800"],
            ["Wide Quad-pol Stripmap (QPSII)", "25 m.", "HH, HV / VV, VH", "71,300", "137,800"],
            ["Wave (WAV)", "10 m.", "HH, HV / VV, VH", "10,700", "14,300"],
            ["Global Observation (GLO)", "500 m.", "HH, HV / VV, VH", "10,700", "14,300"],
            ["Extended Incidence Angle (EXT)", "25 m.", "HH, HV / VV, VH", "42,800", "57,000"],
        ],
        "notes": [],
    },
]

GLOBAL_NOTES = [
    "**ราคาดังกล่าวยังไม่รวมภาษีมูลค่าเพิ่ม",
    "สอบถามรายละเอียดเพิ่มเติมได้ที่ สำนักงานพัฒนาเทคโนโลยีอวกาศและภูมิสารสนเทศ (องค์การมหาชน) "
    "ฝ่ายพัฒนาธุรกิจและการบริการ ศูนย์ราชการเฉลิมพระเกียรติ อาคาร B ชั้น 6 ถนนแจ้งวัฒนะ "
    "แขวงทุ่งสองห้อง เขตหลักสี่ กรุงเทพฯ 10210 โทร 0 2143 9593, 0 2141 4564-66,69 อีเมล usd@gistda.or.th",
    "ที่มา: https://www.gistda.or.th/download/Gistda_Price_List.pdf (ปรับปรุงล่าสุด 11 ต.ค. 2023)",
]


def build_csv(path):
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        for s in SECTIONS:
            w.writerow([s["title"]])
            w.writerow([s["unit"]])
            w.writerow(s["columns"])
            w.writerows(s["rows"])
            for n in s["notes"]:
                w.writerow(["หมายเหตุ: " + n])
            w.writerow([])
        for n in GLOBAL_NOTES:
            w.writerow([n])


def build_xlsx(path):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    head_fill = PatternFill("solid", fgColor="1F4E79")
    head_font = Font(bold=True, color="FFFFFF")

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for s in SECTIONS:
            df = pd.DataFrame(s["rows"], columns=s["columns"])
            df.to_excel(writer, sheet_name=s["sheet"], index=False, startrow=2)
            ws = writer.sheets[s["sheet"]]
            ws.cell(row=1, column=1, value=s["title"]).font = Font(bold=True, size=13)
            ws.cell(row=2, column=1, value=s["unit"]).font = Font(italic=True)
            ncols = len(s["columns"])
            for c in range(1, ncols + 1):
                cell = ws.cell(row=3, column=c)
                cell.fill, cell.font, cell.border = head_fill, head_font, border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for r in range(4, 4 + len(s["rows"])):
                for c in range(1, ncols + 1):
                    ws.cell(row=r, column=c).border = border
            note_row = 4 + len(s["rows"]) + 1
            for i, n in enumerate(s["notes"]):
                cell = ws.cell(row=note_row + i, column=1, value="หมายเหตุ: " + n)
                cell.font = Font(size=9, color="666666")
            for c, col in enumerate(s["columns"], start=1):
                width = max([len(col)] + [len(str(row[c - 1])) for row in s["rows"]])
                ws.column_dimensions[ws.cell(row=3, column=c).column_letter].width = min(width + 4, 55)

        # sheet รวมหมายเหตุท้ายเอกสาร
        ws = writer.book.create_sheet("หมายเหตุ")
        for i, n in enumerate(GLOBAL_NOTES, start=1):
            ws.cell(row=i, column=1, value=n)
        ws.column_dimensions["A"].width = 120


def build_html(path):
    esc = html_mod.escape
    parts = ["""<!DOCTYPE html>
<html lang="th">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>GISTDA Price List — ราคาข้อมูลจากดาวเทียม</title>
<style>
  body { font-family: 'Segoe UI', 'Leelawadee UI', Tahoma, sans-serif; margin: 2rem auto;
         max-width: 1000px; padding: 0 1rem; color: #222; background: #fafafa; }
  h1 { color: #1F4E79; border-bottom: 3px solid #1F4E79; padding-bottom: .4rem; }
  h2 { color: #1F4E79; margin-top: 2.2rem; }
  .unit { font-style: italic; color: #555; margin: .2rem 0 .6rem; }
  table { border-collapse: collapse; width: 100%; background: #fff; margin-bottom: .5rem; }
  th { background: #1F4E79; color: #fff; padding: .5rem .7rem; border: 1px solid #999; }
  td { padding: .4rem .7rem; border: 1px solid #bbb; }
  td.num { text-align: right; white-space: nowrap; }
  tr:nth-child(even) td { background: #eef3f8; }
  ul.notes { font-size: .85rem; color: #666; margin: .3rem 0 0; }
  footer { margin-top: 2.5rem; font-size: .85rem; color: #666; border-top: 1px solid #ccc;
           padding-top: .8rem; }
</style>
</head>
<body>
<h1>GISTDA Price List — ราคาข้อมูลจากดาวเทียม</h1>
"""]
    for s in SECTIONS:
        parts.append(f"<h2>{esc(s['title'])}</h2>")
        parts.append(f"<p class='unit'>{esc(s['unit'])}</p>")
        parts.append("<table><thead><tr>" +
                     "".join(f"<th>{esc(c)}</th>" for c in s["columns"]) +
                     "</tr></thead><tbody>")
        for row in s["rows"]:
            cells = []
            for v in row:
                is_num = v.replace(",", "").replace("N/A", "").strip().isdigit() or v == "N/A"
                cls = " class='num'" if is_num else ""
                cells.append(f"<td{cls}>{esc(v)}</td>")
            parts.append("<tr>" + "".join(cells) + "</tr>")
        parts.append("</tbody></table>")
        if s["notes"]:
            parts.append("<ul class='notes'>" +
                         "".join(f"<li>{esc(n)}</li>" for n in s["notes"]) + "</ul>")
    parts.append("<footer>" +
                 "".join(f"<p>{esc(n)}</p>" for n in GLOBAL_NOTES) +
                 "</footer>\n</body>\n</html>")
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(parts))


if __name__ == "__main__":
    build_csv("Gistda_Price_List.csv")
    build_xlsx("Gistda_Price_List.xlsx")
    build_html("Gistda_Price_List.html")
    total = sum(len(s["rows"]) for s in SECTIONS)
    print(f"done: {len(SECTIONS)} sections, {total} data rows")
